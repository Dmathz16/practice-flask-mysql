from flask import abort, Blueprint, render_template, send_file, current_app, Flask, make_response
from openpyxl import Workbook
from openpyxl.styles import NamedStyle
import io, csv, locale

generate_excel = Blueprint('generate_excel', __name__)

title   = 'Generate Excel'
view    = 'generate-excel'

@generate_excel.route('/generate-excel/')
def index():
    data = { "title": title, "view": view }
    return render_template('generate_excel/index.html', data=data)

@generate_excel.route('/generate-excel/export-excel/')
def export_excel():
    # Create a new Excel workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Current Sheet Title"

    # Add some sample data to the worksheet
    data = [
        ['Name', 'Debit', 'Credit'],
        ['John', 1230, 2425],
        ['Alice', 25, 11],
        ['Bob', 35, 24]
    ]

    for row in data:
        ws.append(row)

    # Apply autofilter to the first column
    ws.auto_filter.ref = ws.dimensions
    
    # Add formulas for debit total and credit total
    ws['B5'] = '=SUM(B2:B4)'
    ws['C5'] = '=SUM(C2:C4)'

    # Format cells to display two decimal places and use commas as thousand separators
    number_style = NamedStyle(name='number_style', number_format='#,##0.00')
    for row in ws.iter_rows(min_row=2, max_row=6, min_col=2, max_col=3):
        for cell in row:
            cell.style = number_style

    # Save the workbook to a BytesIO object
    excel_data = io.BytesIO()
    wb.save(excel_data)
    excel_data.seek(0)

    # Create a response with the Excel file
    response = make_response(excel_data.getvalue())
    response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    response.headers['Content-Disposition'] = 'attachment; filename=sample.xlsx'

    return response